"""Lee la lista de CUFEs desde Excel (.xlsx) o CSV y la normaliza a Jobs.

Detecta columnas sin importar mayusculas/acentos. Columna obligatoria: CUFE
(acepta encabezados como 'CUFE', 'CUFE/CUDE', 'CUDE', 'documentkey'...).
Columnas opcionales aprovechadas si existen: tipo de documento, prefijo, folio,
nombre emisor (para nombrar y organizar la salida).

Nota sobre la contrasena: la clave de los documentos es el NIT del USUARIO que
consulta (global, DIAN_NIT), no el NIT del emisor de cada fila. Por eso NO se
mapea 'NIT Emisor' como contrasena por fila.
"""
from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class Job:
    cufe: str
    nit: str | None = None       # sobrescribe el NIT global solo si la fila trae 'nit'
    nombre: str | None = None
    proveedor: str | None = None
    tipo: str | None = None
    prefijo: str | None = None
    folio: str | None = None
    row: int = 0
    warnings: list[str] = field(default_factory=list)

    def safe_name(self) -> str:
        """Nombre de archivo: 'nombre', o 'PREFIJO-FOLIO', o cola del CUFE."""
        if self.nombre:
            base = self.nombre
        elif self.prefijo and self.folio:
            base = f"{self.prefijo}-{self.folio}"
        elif self.folio:
            base = self.folio
        else:
            base = f"factura_{self.cufe[-12:]}"
        base = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", base).strip()
        return base or self.cufe[-12:]

    def folder(self) -> str:
        """Subcarpeta de salida: proveedor saneado, o vacio."""
        if not self.proveedor:
            return ""
        return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", self.proveedor).strip()


def _norm(text: str) -> str:
    text = unicodedata.normalize("NFKD", text or "")
    text = "".join(c for c in text if not unicodedata.combining(c))
    return text.strip().lower()


# Coincidencia exacta (tras normalizar) -> campo canonico.
_HEADER_EXACT = {
    "cufe": "cufe", "cude": "cufe", "cufe/cude": "cufe", "cufe cude": "cufe",
    "documentkey": "cufe", "clave": "cufe", "codigo": "cufe", "codigo unico": "cufe",
    "nit": "nit", "password": "nit", "contrasena": "nit",
    "nombre": "nombre", "archivo": "nombre", "nombre archivo": "nombre",
    "proveedor": "proveedor", "nombre emisor": "proveedor", "emisor": "proveedor",
    "tipo de documento": "tipo", "tipo": "tipo",
    "prefijo": "prefijo",
    "folio": "folio", "numero": "folio", "no": "folio",
}

_CUFE_RE = re.compile(r"^[0-9a-fA-F]{40,120}$")


def _map_headers(headers: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for i, h in enumerate(headers):
        key = _norm(h)
        field_name = _HEADER_EXACT.get(key)
        if not field_name and ("cufe" in key or "cude" in key):
            field_name = "cufe"  # tolera variantes: 'CUFE/CUDE', 'cufe_uuid'...
        if field_name and field_name not in mapping.values():
            mapping[i] = field_name
    return mapping


def _rows_from_csv(path: Path) -> list[list[str]]:
    sample = path.read_text(encoding="utf-8-sig", errors="replace")[:4096]
    delim = ";" if sample.count(";") > sample.count(",") else ","
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [row for row in csv.reader(fh, delimiter=delim)]


def _rows_from_xlsx(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [["" if c is None else str(c) for c in r] for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def read_jobs(path: str | Path) -> list[Job]:
    path = Path(path).expanduser()
    if not path.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {path}")

    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        rows = _rows_from_xlsx(path)
    elif suffix in {".csv", ".txt", ""}:
        rows = _rows_from_csv(path)
    else:
        raise ValueError(f"Formato no soportado: {suffix}. Usa .xlsx o .csv")

    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        raise ValueError(f"El archivo {path.name} no tiene datos.")

    header_map = _map_headers(rows[0])
    if "cufe" not in header_map.values():
        raise ValueError(
            "No encontre una columna de CUFE. Encabezados aceptados: cufe, "
            "cufe/cude, cude, documentkey, clave, codigo. "
            f"Detectados: {rows[0]}"
        )

    jobs: list[Job] = []
    seen: set[str] = set()
    for idx, row in enumerate(rows[1:], start=2):
        rec: dict[str, str] = {}
        for col, name in header_map.items():
            if col < len(row):
                rec[name] = (row[col] or "").strip()

        cufe = rec.get("cufe", "")
        if not cufe or cufe in seen:
            continue
        seen.add(cufe)

        job = Job(
            cufe=cufe,
            nit=rec.get("nit") or None,
            nombre=rec.get("nombre") or None,
            proveedor=rec.get("proveedor") or None,
            tipo=rec.get("tipo") or None,
            prefijo=rec.get("prefijo") or None,
            folio=rec.get("folio") or None,
            row=idx,
        )
        if not _CUFE_RE.match(cufe):
            job.warnings.append(
                f"Fila {idx}: el CUFE no parece hex valido (len={len(cufe)}). Se intentara igual."
            )
        jobs.append(job)

    if not jobs:
        raise ValueError("No se encontraron CUFEs validos en el archivo.")
    return jobs
